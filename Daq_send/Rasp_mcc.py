# -*- coding: utf-8 -*-
"""
Created on Fri Feb 25 13:44:38 2022

@author: vollmera
"""
import os

from ReadList import xls_handling


from openpyxl.styles import Color, PatternFill, Font, Border
from openpyxl.styles import colors
redFill = PatternFill(start_color='FFFF0000',
                   end_color='FFFF0000',
                   fill_type='solid')

greenFill = PatternFill(start_color='0099CC00',
                   end_color='0099CC00',
                   fill_type='solid')


folder=os.getcwd()
file="Raspi_confgi.xlsx"
