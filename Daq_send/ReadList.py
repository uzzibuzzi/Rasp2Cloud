# -*- coding: utf-8 -*-
"""
importer for a configuration file for MCC daqhats 

@author: vollmera
"""
import pandas as pd
import numpy as np
import os
import openpyxl




from openpyxl import load_workbook


class xls_handling:
    def __init__(self):
        self.socklist = []
        self.wb=0
        self.ws=0
        self.stocklist=[]
        
    def load_list_from_xls(self,filename,sheetname):
        self.wb= load_workbook(filename)
        self.ws= self.wb[sheetname]
        wknRow=1
        
        for row in self.ws.iter_rows(min_row=2, max_col=wknRow, max_row=5):
            for cell in row:
                self.stocklist.append(cell.value)

    def getSheets(self):
        """return list of sheetnames"""
        return self.wb.sheetnames
    
    def getKeys(self,axis):
        newList=[]
        columEnd=1
        rowEnd=1
        if axis == 1:
            columEnd=100
            startRow=1
            #print("columEnd")
        if axis==0:
            rowEnd=100
            startRow=2
            #print("row end")
            
        for col in self.ws.iter_cols(min_row=startRow, max_col=columEnd, max_row=rowEnd):
            for cell in col:
                if cell.value is None:
                     return newList
                newList.append(cell.value)
        print("End of counter reached")
        return newList
  

            
    def get_colum_From(self,label):
        aaa=self.getKeys(1)
        column=aaa.index(label)  
        rowList=[]
        for ROWS  in self.ws.iter_rows(2, 1+len(self.getKeys(0))):
            rowList.append(ROWS[column].value)
        return rowList
               
    def get_row_From(self,label):
        aaa=self.getKeys(0)
        rowselect=aaa.index(label)+1 
        ColList=[]
        for COLUM  in self.ws.iter_cols(1, len(self.getKeys(1))):
            ColList.append(COLUM[rowselect].value)
        return ColList

    def NodeName(self):
        nameOfNode=self.get_colum_From("Name of node")
        numberOfNode=self.get_colum_From("Nodes used")    
        self.nodeDict={}
        
        for each in range(len(numberOfNode)):
            new={str(numberOfNode[each]):nameOfNode[each]}
            self.nodeDict.update(new)
        return self.nodeDict
    def configRasp(self):
        
        print("setup for rasp Nr. {}   file version of schema {}".format(self.get_colum_From("Raspberry ID")[0],self.get_colum_From("Config schema")[0]))            
        for each in range(1,1+len(self.NodeName())):
            print("Config node {} as mcc device {} ".format(each,self.nodeDict.get(str(each))))
            



if __name__ == "__main__":
    file="Raspi_confgi.xlsx"

    abc=xls_handling()
    abc.load_list_from_xls(file,"Config_overview")
    allSheets=abc.getSheets() 
    abc.configRasp()


# check if mcc config is known
mccKnown=[888,155]

#make all signal konig

for each in range(1,len(allSheets)):
    abc.load_list_from_xls(file,allSheets[each])
    print("node nr. {} mcc found as {} ".format(each,allSheets[each]))
    keyList=abc.getKeys(1)
    for each in keyList:
        print("\t{} {}".format(each,abc.get_colum_From(each)))


#abc[1] each in mccKnown 



"""
make an summary side to use in cloud
   # allSheets.append("all") 
   """
